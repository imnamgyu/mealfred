#!/usr/bin/env python3
"""
seed-institutions.py — 공공데이터 표준데이터(어린이집·유치원·학교) → public.institutions 적재.

사용:  python3 scripts/seed-institutions.py <파일.csv|xlsx> <type>
  type = daycare(어린이집) | kindergarten(유치원) | school(초중고)
  예)  python3 scripts/seed-institutions.py "~/Downloads/일반 현황.xlsx" kindergarten

특징:
  - CSV 인코딩 자동(cp949/euc-kr/utf-8-sig), xlsx 지원(openpyxl).
  - 제목/빈 행을 건너뛰고 '진짜 헤더 행'(이름+주소 포함)을 자동 탐지(유치원 파일은 1~2행이 제목).
  - 표준데이터 헤더를 부분일치로 매핑 + 원장명(있으면) + 주소에서 시도/시군구/동 파싱.
  - (type, name_norm, address) 자연키 in-script dedup 후 Supabase REST upsert(merge-duplicates) 배치.
  - 환경: web/.env.local 의 SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY 자동 로드.
"""
import sys, os, csv, json, re, urllib.request

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def load_env():
    env = {}
    p = os.path.join(ROOT, ".env.local")
    if os.path.exists(p):
        for line in open(p, encoding="utf-8"):
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip().strip('"').strip("'")
    return env

def norm_name(s):
    return re.sub(r"\s+", "", (s or "")).strip()

def parse_dong(addr):
    if not addr:
        return None
    m = re.search(r"([가-힣]+(?:[0-9]*가)?[동읍면])(?:\s|$|[0-9])", addr)
    return m.group(1) if m else None

SIDO_RE = re.compile(
    r"^(서울특별시|부산광역시|대구광역시|인천광역시|광주광역시|대전광역시|울산광역시|세종특별자치시|"
    r"경기도|강원특별자치도|강원도|충청북도|충청남도|전북특별자치도|전라북도|전라남도|"
    r"경상북도|경상남도|제주특별자치도|충북|충남|전북|전남|경북|경남|서울|부산|대구|인천|광주|대전|울산|세종)"
)
def parse_sido_sigungu(addr):
    """주소 문자열 → (시도, 시군구). 별도 컬럼이 없을 때(유치원) 사용."""
    if not addr:
        return None, None
    a = addr.strip()
    m = SIDO_RE.match(a)
    sido = m.group(1) if m else (a.split()[0] if a.split() else None)
    rest = a[len(sido):].strip() if (sido and a.startswith(sido)) else a
    m2 = re.search(r"([가-힣]+(?:특별자치시|시|군|구))", rest)
    return sido, (m2.group(1) if m2 else None)

def to_float(v):
    try:
        return float(str(v).strip())
    except Exception:
        return None

def to_int(v):
    try:
        return int(float(str(v).strip()))
    except Exception:
        return None

# 헤더 부분일치 후보(표준데이터 연도별 미세변동 흡수). 앞에서부터 먼저 맞는 컬럼 사용.
FIELD_HINTS = {
    "name":      ["어린이집명", "유치원명", "학교명", "시설명", "기관명"],
    "inst_type": ["유형구분", "설립유형", "어린이집유형", "유형", "설립"],
    "status":    ["운영현황", "영업상태", "상태"],
    "sido":      ["시도", "시·도", "광역시도"],
    "sigungu":   ["시군구", "시·군·구"],
    "address":   ["도로명주소", "소재지도로명", "주소", "소재지"],
    "zipcode":   ["우편번호"],
    "capacity":  ["총정원", "인가총정원", "정원수", "정원"],
    "lat":       ["위도", "ycrdnt", "latitude"],
    "lng":       ["경도", "xcrdnt", "longitude"],
    "ext_code":  ["어린이집코드", "유치원코드", "시설코드", "표준학교코드", "학교코드"],
}
NAME_HEADER_KEYS = ["유치원명", "어린이집명", "학교명", "시설명", "기관명"]

def build_colmap(headers):
    cmap = {}
    low = [(h, (h or "").replace(" ", "").lower()) for h in headers]
    for field, hints in FIELD_HINTS.items():
        for hint in hints:
            hl = hint.replace(" ", "").lower()
            hit = next((h for h, hlow in low if hl in hlow), None)
            if hit:
                cmap[field] = hit
                break
    return cmap

def _all_rows(path):
    """파일 → list[list[str]] (모든 셀). xlsx=openpyxl · xls=xlrd · csv=인코딩자동."""
    pl = path.lower()
    if pl.endswith(".xlsx"):
        import warnings; warnings.filterwarnings("ignore")
        from openpyxl import load_workbook
        ws = load_workbook(path, data_only=True).active   # read_only X (차원 정확)
        return [["" if c is None else str(c).strip() for c in row]
                for row in ws.iter_rows(values_only=True)]
    if pl.endswith(".xls"):
        import warnings; warnings.filterwarnings("ignore")
        import xlrd
        ws = xlrd.open_workbook(path).sheet_by_index(0)
        return [[str(ws.cell_value(i, j)).strip() for j in range(ws.ncols)] for i in range(ws.nrows)]
    raw = open(path, "rb").read()
    text = None
    for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
        try:
            text = raw.decode(enc); break
        except UnicodeDecodeError:
            continue
    if text is None:
        sys.exit("CSV 인코딩 판별 실패")
    return [row for row in csv.reader(text.splitlines())]

def read_rows(path):
    rows = _all_rows(path)
    hdr_idx = None
    for i, row in enumerate(rows[:25]):   # 제목/빈 행 스킵 — 이름+주소가 같이 있는 첫 행이 헤더
        cells = [(c or "").replace(" ", "") for c in row]
        if any(any(k in c for k in NAME_HEADER_KEYS) for c in cells) and any(("주소" in c or "소재지" in c) for c in cells):
            hdr_idx = i; break
    if hdr_idx is None:
        sys.exit(f"헤더 행(이름+주소 컬럼)을 못 찾음. 상위행={rows[:3]}")
    headers = rows[hdr_idx]
    print(f"[header] row{hdr_idx + 1}: {headers}")
    for row in rows[hdr_idx + 1:]:
        if all((c or "").strip() == "" for c in row):
            continue
        yield dict(zip(headers, row))

def main():
    if len(sys.argv) < 3:
        sys.exit(__doc__)
    path, itype = os.path.expanduser(sys.argv[1]), sys.argv[2]
    if itype not in ("daycare", "kindergarten", "school"):
        sys.exit("type 은 daycare|kindergarten|school")
    # 옵션: --exclude-sido 제주  (해당 시도 제외)
    rest_args = sys.argv[3:]
    exclude_sido = None
    for i, a in enumerate(rest_args):
        if a == "--exclude-sido" and i + 1 < len(rest_args):
            exclude_sido = rest_args[i + 1]
        elif a.startswith("--exclude-sido="):
            exclude_sido = a.split("=", 1)[1]
    env = load_env()
    base = (env.get("SUPABASE_URL") or env.get("NEXT_PUBLIC_SUPABASE_URL") or "").rstrip("/")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY") or ""
    if not base or not key:
        sys.exit("SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY 누락(.env.local)")

    rows = read_rows(path)
    first = next(rows, None)
    if first is None:
        sys.exit("빈 파일")
    cmap = build_colmap(list(first.keys()))
    if "name" not in cmap or "address" not in cmap:
        sys.exit(f"이름/주소 컬럼 못 찾음.\n헤더={list(first.keys())}\n매핑={cmap}")
    print(f"[map] {cmap}")

    src = {"daycare": "childcare_std", "kindergarten": "kindergarten_std", "school": "neis"}[itype]
    seen, recs = set(), []
    skipped = {"closed": 0, "sido": 0}
    def add(row):
        name = (row.get(cmap["name"]) or "").strip()
        addr = (row.get(cmap.get("address", "")) or "").strip()
        if not name:
            return
        status = (row.get(cmap.get("status", "")) or "").strip()
        if any(x in status for x in ("폐지", "폐원", "휴지", "휴원")) or re.search(r"폐원|폐지|휴원", name):   # 폐지·휴지 제외(운영 중만)
            skipped["closed"] += 1; return
        p_sido, p_sigungu = parse_sido_sigungu(addr)
        sido_v = (row.get(cmap.get("sido", "")) or "").strip() or p_sido
        if exclude_sido and sido_v and exclude_sido in sido_v:   # 제주 등 시도 제외
            skipped["sido"] += 1; return
        nn = norm_name(name)
        k = (itype, nn, addr)
        if k in seen:
            return
        seen.add(k)
        rec = {
            "type": itype, "name": name, "name_norm": nn,
            "inst_type": (row.get(cmap.get("inst_type", "")) or "").strip() or None,
            "sido": sido_v,
            "sigungu": (row.get(cmap.get("sigungu", "")) or "").strip() or p_sigungu,
            "dong": parse_dong(addr),
            "address": addr or None,
            "zipcode": (row.get(cmap.get("zipcode", "")) or "").strip() or None,
            "capacity": to_int(row.get(cmap.get("capacity", ""))),
            "lat": to_float(row.get(cmap.get("lat", ""))),
            "lng": to_float(row.get(cmap.get("lng", ""))),
            "ext_code": (row.get(cmap.get("ext_code", "")) or "").strip() or None,
            "source": src,
        }
        recs.append(rec)   # 원장명 미사용(이사님) — 기관명+주소로 분별
    add(first)
    for r in rows:
        add(r)
    print(f"[rows] 적재대상 {len(recs)}건 · 제외(폐지/휴원 {skipped['closed']} · 시도 {skipped['sido']}) · 예: {recs[0]['name']}·{recs[0].get('director')}·{recs[0].get('sido')} {recs[0].get('sigungu')}")

    url = f"{base}/rest/v1/institutions?on_conflict=type,name_norm,address"
    H = {"apikey": key, "Authorization": "Bearer " + key, "Content-Type": "application/json",
         "Prefer": "resolution=merge-duplicates,return=minimal"}
    ok = 0
    for i in range(0, len(recs), 500):
        batch = recs[i:i + 500]
        req = urllib.request.Request(url, data=json.dumps(batch).encode("utf-8"), headers=H, method="POST")
        try:
            with urllib.request.urlopen(req) as resp:
                ok += len(batch); print(f"  upsert {ok}/{len(recs)}")
        except urllib.error.HTTPError as e:
            sys.exit(f"upsert 실패 {e.code}: {e.read().decode()[:300]}")
    print(f"[done] {ok}건 적재(type={itype})")

if __name__ == "__main__":
    main()
