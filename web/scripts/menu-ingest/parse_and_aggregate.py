#!/usr/bin/env python3
"""
영유아 식단 적재 파이프라인 — 1단계: 파싱 + 정제 + 빈도집계 + 2축 도감 재정의안 산출

입력:  --in <폴더>   (어린이급식관리지원센터 식단 파일들: .xlsx / .csv / .zip)
출력:  <out>/dogam_redefine.csv   ← 2축(영유아×초등) 도감 재등급 + 승격후보
       <out>/ing_freq.csv         ← 식재료별 영유아 등장률
       <out>/dish_decomp.csv      ← 음식→식재료 분해
       <out>/promo_candidates.json← 승격후보(다음 단계 안전스크린 워크플로 입력)

핵심:
 - **헤더 자동감지**: 시트에서 '음식명/메뉴명' '식재료명/재료명' 컬럼을 찾아 정렬 → 컬럼 어긋남 노이즈 해결
 - bare-XLSX(발주량산출서)와 zip-of-xlsx 둘 다 처리
 - 단위 = 파일 1개(센터-월 가정). 등장률 = 그 식재료가 등장한 파일 비율
 - 노이즈 컷: 숫자/날짜/알레르기기호(①②)/브랜드 → 식재료칸에서 제거
 - 영유아축 = 본 수집, 초등축 = SQL ingredients.elem_count(NEIS)

실행:  python parse_and_aggregate.py --in ~/Desktop/menus --out /tmp/menu_out
사용법 전체는 README.md 참조.
"""
import sys, os, re, io, json, csv, zipfile, argparse, urllib.request, collections
try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl 필요: pip install openpyxl")

# ── 식재료 표준화 / 양념 제외 ──────────────────────────────────────────────
CANON = {'쌀':'멥쌀','백미':'멥쌀','쌀밥':'멥쌀','멥쌀밥':'멥쌀','누룽지':'멥쌀','잡곡':'보리',
    '콩':'콩(대두)','대두':'콩(대두)','떡':'멥쌀떡','요거트':'요구르트','달걀':'계란','쇠고기':'소고기',
    '보리차':'보리','옥수수차':'옥수수','큰느타리버섯':'새송이버섯',
    '대멸치':'멸치','중멸치':'멸치','잔멸치':'멸치','지리멸':'멸치','국물멸치':'멸치'}
SEASON = set('소금 물 설탕 흑설탕 간장 진간장 국간장 양조간장 조선간장 재래 참기름 들기름 식용유 콩기름 '
    '카놀라유 올리브유 포도씨유 후추 후춧가루 통후추 고춧가루 깨소금 참깨 들깨 통깨 깨 마늘 다진마늘 생강 '
    '식초 맛술 미림 물엿 조청 올리고당 꿀 케첩 케찹 마요네즈 다시다 미원 조미료 멸치액젓 까나리액젓 액젓 '
    '새우젓 전분 감자전분 옥수수전분 밀가루 부침가루 튀김가루 빵가루 카레가루 청주 맛소금 베이킹파우더 식소다 '
    '버터 마가린 된장 고추장 쌈장 춘장 청국장 막장 토마토케첩 데리야끼 유자청 매실청 청양고추 홍고추 풋고추 '
    '식염 천일염 정제소금 맛간장 액상과당 향신료'.split())

def canon(nm):
    nm = nm.strip().lstrip('(').strip()
    if nm in CANON: return CANON[nm]
    return re.split(r'[_,(]', nm)[0].strip()

def is_season(c):
    return c in SEASON or '소스' in c or '액젓' in c or '페이스트' in c or '드레싱' in c

ALLERGEN = '①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳'
def strip_allergen(s):                          # 음식명 끝의 알레르기 번호 제거
    return re.sub(r'[%s\s]+$' % ALLERGEN, '', s).strip()

def is_noise(s):                                # 식재료칸에 샌 비식재료
    if not re.search(r'[가-힣]', s): return True            # 숫자/영문 파편
    if re.fullmatch(r'[\d./\s]+', s): return True
    if re.search(r'[%s]' % ALLERGEN, s): return True        # 알레르기기호 = 음식명 잔재
    if re.search(r'\d+\s*\[[월화수목금토일]\]', s): return True  # 날짜
    if 'kcal' in s or s in ('에너지','팽창제','조미소','산'): return True
    return False

# ── 헤더 자동감지: '음식명/식재료명' 컬럼 위치 찾기 ───────────────────────────
DISH_H = ('음식명','메뉴명','요리명')
ING_H  = ('식재료명','재료명','식품명')
SLOT_H = ('구분','끼니','식사')

def detect_cols(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 20 or not row: continue
        cells = [str(c).strip() if c is not None else '' for c in row]
        di = next((j for j,c in enumerate(cells) if c in DISH_H), None)
        ii = next((j for j,c in enumerate(cells) if c in ING_H), None)
        if di is not None and ii is not None:
            si = next((j for j,c in enumerate(cells) if c in SLOT_H), None)
            return {'dish':di,'ing':ii,'slot':si,'header_row':i}
    return None

def parse_xlsx(data):
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out = []
    for sn in wb.sheetnames:
        if sn.strip() in ('표지','목차'): continue
        ws = wb[sn]
        cols = detect_cols(ws)
        if not cols:                            # 헤더 못찾으면 표준 위치 가정(날짜|구분|음식|식재료)
            cols = {'dish':2,'ing':3,'slot':1,'header_row':-1}
        dish = slot = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i <= cols['header_row'] or not row: continue
            def cell(k):
                j = cols[k]
                return row[j] if (j is not None and j < len(row)) else None
            dn, ing = cell('dish'), cell('ing')
            s = cell('slot')
            if isinstance(dn, str) and dn.strip() and dn.strip() not in DISH_H:
                dish = strip_allergen(dn.strip())
            if isinstance(s, str) and s.strip() and s.strip() not in SLOT_H:
                slot = s.strip()
            if isinstance(ing, str) and ing.strip() and ing.strip() not in ING_H:
                c = canon(ing)
                if c and dish and not is_noise(c):
                    out.append((dish, c, slot))
    return out

def parse_file(path):
    name = os.path.basename(path)
    if name.lower().endswith('.csv'):
        rows = []
        with open(path, encoding='utf-8', errors='replace') as f:
            for r in csv.reader(f):
                if len(r) >= 4:
                    d, ing = r[2].strip(), canon(r[3])
                    if d and ing and not is_noise(ing): rows.append((d, ing, r[1] if len(r)>1 else None))
        return rows
    data = open(path, 'rb').read()
    if data[:2] != b'PK': return []
    names = zipfile.ZipFile(io.BytesIO(data)).namelist()
    if 'xl/workbook.xml' in names:              # bare xlsx
        return parse_xlsx(data)
    out = []                                     # zip-of-xlsx
    zf = zipfile.ZipFile(io.BytesIO(data))
    for info in zf.infolist():
        if info.filename.lower().endswith('.xlsx'):
            try: out += parse_xlsx(zf.read(info))
            except Exception: pass
    return out

# ── SQL ingredients (초등축 elem_count + 현 도감 등급) ─────────────────────────
def load_env():
    web = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
    env = {}
    for line in open(os.path.join(web, '.env.local')):
        m = re.match(r'\s*([A-Z_]+)\s*=\s*(.+?)\s*$', line)
        if m: env.setdefault(m.group(1), m.group(2).strip().strip('"'))
    return env['NEXT_PUBLIC_SUPABASE_URL'], env['SUPABASE_SERVICE_ROLE_KEY']

def sql_ingredients():
    URL, KEY = load_env()
    H = {'apikey': KEY, 'Authorization': f'Bearer {KEY}'}
    out, off = [], 0
    sel = 'ingredients?select=name,slug,grade_label,grade_star,elem_count,food_group,allergens,warning,meta'
    while True:
        d = json.loads(urllib.request.urlopen(urllib.request.Request(
            f'{URL}/rest/v1/{sel}&offset={off}&limit=1000', headers=H), timeout=40).read())
        out += d
        if len(d) < 1000: break
        off += 1000
    return out

def yu_grade(p): return '매일군' if p>=.7 else ('자주군' if p>=.35 else ('가끔군' if p>=.1 else '드물군'))
def el_grade(c): return '초등매일' if c>=300 else ('초등자주' if c>=100 else ('초등가끔' if c>=10 else '초등드묾'))

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--in', dest='indir', required=True, help='식단 파일 폴더')
    ap.add_argument('--out', dest='outdir', default='/tmp/menu_out')
    a = ap.parse_args()
    os.makedirs(a.outdir, exist_ok=True)

    files = [os.path.join(a.indir, f) for f in sorted(os.listdir(a.indir))
             if f.lower().endswith(('.xlsx', '.csv', '.zip'))]
    print(f'입력 파일 {len(files)}개')
    unit_ings = collections.defaultdict(set)   # 파일단위 등장
    dish_ings = collections.defaultdict(set)
    noise_n = 0
    for fp in files:
        rows = parse_file(fp)
        u = os.path.basename(fp)
        for dish, ing, slot in rows:
            unit_ings[ing].add(u); dish_ings[dish].add(ing)
        print(f'  {os.path.basename(fp)[:40]:40s} {len(rows)}행')
    TOT = len(files)
    print(f'\n단위(파일) {TOT} · 식재료 {len(unit_ings)} · 음식 {len(dish_ings)}')

    sql = sql_ingredients()
    byname = {}
    for x in sql:
        byname[x['name']] = x
        if x.get('slug'): byname.setdefault(x['slug'], x)
    def tier(x): return (x.get('meta') or {}).get('tier', '-')

    # ── 2축 도감 재정의안 ──
    RANK={'드물군':0,'가끔군':1,'자주군':2,'매일군':3}; ELR={'초등드묾':0,'초등가끔':1,'초등자주':2,'초등매일':3}
    TIERN=['드물군','가끔군','자주군','매일군']
    dogam_rows=[]; promo=[]
    for ing, units in sorted(unit_ings.items(), key=lambda kv:-len(kv[1])):
        pct = len(units)/TOT if TOT else 0
        x = byname.get(ing) or byname.get(canon(ing))
        is_dogam = bool(x) and tier(x) in ('-','dogam') and (x.get('grade_label') or '')!='향신료'
        if x and is_dogam:
            yg, el = yu_grade(pct), (x.get('elem_count') or 0)
            eg = el_grade(el); comb = TIERN[max(RANK[yg], ELR[eg])]
            yr,er = RANK[yg], ELR[eg]
            flag = '공통' if (yr>=2 and er>=2) else ('영유아특이' if yr-er>=2 else ('초등특이' if er-yr>=2 else '공통'))
            cur = x.get('grade_label') or '(없음)'
            curlvl = {'필수':3,'권장':2}.get(cur)
            act = ('신규등급='+comb) if cur=='(없음)' else (
                  '유지' if curlvl is None else (
                  '상향→'+comb if TIERN.index(comb)>curlvl else (
                  '하향검토→'+comb if TIERN.index(comb)<curlvl else '유지')))
            dogam_rows.append([ing, x.get('food_group') or '', round(pct*100,1), yg, el, eg, comb, flag, cur, act])
        elif pct >= 0.10 and not is_dogam:      # 승격후보(비도감·영유아 가끔군+)
            promo.append({'ing':ing,'yu_pct':round(pct*100,1),'yu_dishes':len(dish_ings.get(ing,[])),'in_sql':bool(x)})

    with open(os.path.join(a.outdir,'dogam_redefine.csv'),'w',newline='') as f:
        w=csv.writer(f)
        w.writerow([f'도감 재정의안 — 2축(영유아 {TOT}파일 × 초등 NEIS)'])
        w.writerow(['식재료','식품군','영유아등장률%','영유아등급','초등count','초등등급','결합등급','연령플래그','현등급','액션'])
        for r in dogam_rows: w.writerow(r)
    with open(os.path.join(a.outdir,'ing_freq.csv'),'w',newline='') as f:
        w=csv.writer(f); w.writerow(['식재료','영유아등장률%','등장파일수','음식수','도감여부'])
        for ing,units in sorted(unit_ings.items(),key=lambda kv:-len(kv[1])):
            x=byname.get(ing); dg=bool(x) and tier(x) in ('-','dogam')
            w.writerow([ing,round(len(units)/TOT*100,1),len(units),len(dish_ings.get(ing,[])),'도감' if dg else ('유니버스' if x else '신규')])
    with open(os.path.join(a.outdir,'dish_decomp.csv'),'w',newline='') as f:
        w=csv.writer(f); w.writerow(['음식명','식재료','식재료수'])
        for d,igs in sorted(dish_ings.items()): w.writerow([d,', '.join(sorted(igs)),len(igs)])
    json.dump({'TOT':TOT,'candidates':promo}, open(os.path.join(a.outdir,'promo_candidates.json'),'w'), ensure_ascii=False)

    print(f'\n산출 → {a.outdir}/')
    print(f'  dogam_redefine.csv  ({len(dogam_rows)} 도감행)')
    print(f'  promo_candidates.json ({len(promo)} 승격후보 → 다음: 안전스크린 워크플로)')
    print('\n다음 단계: 워크플로 dogam-promotion-safety-screen 에 promo_candidates.json 투입 → 승격확정')
    print('그 다음: python apply_dogam_redefine.py --redefine dogam_redefine.csv --promoted screen_result.json')

if __name__ == '__main__':
    main()
